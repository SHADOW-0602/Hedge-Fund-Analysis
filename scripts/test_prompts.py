
import os
import requests
import json
import time
import datetime
from datetime import date
from dotenv import load_dotenv
import yfinance as yf
import re
import pandas as pd
import random
import textwrap
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Load environment variables
load_dotenv()

# Configuration
TICKERS = ["TSLA", "GOOGL", "AMZN", "NVDA", "AAPL", "MSFT", "META"] 
STRATEGIES = ["Detailed", "Crisp", "PriceContext"]

def get_gemini_keys():
    """Load all available Gemini API keys from environment"""
    keys = []
    
    # Check primary key
    key1 = os.getenv('GEMINI_API_KEY')
    if key1:
        keys.append((key1, "GEMINI_API_KEY"))
        
    # Check numbered keys 2-6
    for i in range(2, 7):
        key_name = f'GEMINI_API_KEY_{i}'
        key_val = os.getenv(key_name)
        if key_val:
            keys.append((key_val, key_name))
            
    return keys

def fetch_quote_data(ticker):
    """Fetch real-time quote data using yfinance"""
    try:
        stock = yf.Ticker(ticker)
        info = stock.fast_info
        last_price = info.last_price
        prev_close = info.previous_close
        
        if prev_close and prev_close > 0:
            change = last_price - prev_close
            change_percent = (change / prev_close) * 100
            return {
                'ticker': ticker,
                'price': round(last_price, 2),
                'change_percent': round(change_percent, 2)
            }
    except Exception as e:
        print(f"Error fetching quote for {ticker}: {e}")
    return None

def fetch_news_for_ticker(ticker):
    """Fetch news from multiple sources"""
    all_news = []
    print(f"Fetching news for {ticker}...")

    # 1. NewsAPI
    try:
        newsapi_key = os.getenv('NEWSAPI_KEY')
        if newsapi_key:
            url = f"https://newsapi.org/v2/everything?q={ticker}&apiKey={newsapi_key}&pageSize=5&sortBy=publishedAt"
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                data = response.json()
                for article in data.get('articles', []):
                    all_news.append({
                        'title': article.get('title', ''),
                        'source': article.get('source', {}).get('name', 'NewsAPI'),
                        'description': article.get('description', '')
                    })
    except Exception as e:
        pass

    # 2. Polygon
    try:
        polygon_key = os.getenv('POLYGON_API_KEY')
        if polygon_key:
            url = f"https://api.polygon.io/v2/reference/news?ticker={ticker}&limit=5&apiKey={polygon_key}"
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                data = response.json()
                for article in data.get('results', []):
                    all_news.append({
                        'title': article.get('title', ''),
                        'source': article.get('publisher', {}).get('name', 'Polygon'),
                        'description': article.get('description', '')
                    })
    except Exception as e:
        pass

    # 3. Yahoo Finance
    try:
        stock = yf.Ticker(ticker)
        news = stock.news
        if news:
            for article in news[:5]:
                all_news.append({
                    'title': article.get('title', ''),
                    'source': article.get('publisher', 'Yahoo Finance'),
                    'description': article.get('title', '')
                })
    except Exception as e:
        pass

    print(f"  Found {len(all_news)} articles total.")
    return all_news

def generate_summary(ticker, strategy, news_articles, keys):
    """Generate summary using Gemini 2.5 Flash"""
    if not news_articles:
        return "No news found."

    # Prepare news text
    news_text = f"Stock Ticker: {ticker}\n\n"
    for idx, article in enumerate(news_articles[:10], 1):
        news_text += f"{idx}. {article['title']}\n"
        news_text += f"   Source: {article['source']}\n"
        news_text += f"   {article['description']}\n\n"

    # Define Strategy Instruction
    if strategy == "PriceContext":
        quote_data = fetch_quote_data(ticker)
        change_str = "N/A"
        if quote_data:
            sign = "+" if quote_data['change_percent'] >= 0 else ""
            change_str = f"{sign}{quote_data['change_percent']}%"
        content_instruction = f"Today's Price change of {ticker} is {change_str}. Prioritize the most recent news articles and give a crisp, relevant accurate summary that can partly explain the price changes."
        
    elif strategy == "Crisp":
        content_instruction = f"Give a crisp, relevant and accurate summary of the latest developments for {ticker} with a clear what, why and how."
        
    else: # Detailed
        content_instruction = f"Analyze the following news articles about {ticker} stock and create a **detailed and comprehensive** summary."

    prompt = f"""{content_instruction}
    
    IMPORTANT: Format the executive summary as a **list of bullet points**.
    RETURN ONLY JSON with key "executive_summary".
    
{news_text}

    Please provide a JSON response with the following structure:
    {{
        "executive_summary": "<ul><li>Key point 1...</li><li>Key point 2...</li></ul>"
    }}
"""

    payload = {
        "system_instruction": {
            "parts": [{"text": "You are a financial analyst AI. You provide detailed stock news summaries in structured JSON format with HTML content."}]
        },
        "contents": [
            {"role": "user", "parts": [{"text": prompt}]}
        ],
        "generationConfig": {
            "response_mime_type": "application/json",
            "temperature": 0.3
        }
    }

    # Rotate Keys
    available_keys = list(keys)
    random.shuffle(available_keys)

    for attempt, (api_key, key_name) in enumerate(available_keys):
        # Using Gemini 2.5 Flash as requested
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        headers = {'Content-Type': 'application/json'}
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=25)
            
            if response.status_code == 429:
                print(f"Key {key_name} Rate Limited. Waiting 5s.")
                time.sleep(5)
                continue
            
            if response.status_code == 200:
                result = response.json()
                try:
                    candidates = result.get('candidates', [])
                    if candidates:
                        content_text = candidates[0].get('content', {}).get('parts', [{}])[0].get('text', '')
                        if content_text:
                            clean_text = content_text.replace('```json', '').replace('```', '').strip()
                            json_content = json.loads(clean_text)
                            return json_content.get('executive_summary', 'No summary key found.')
                except:
                    pass
                # If parsing fails, try next key or return valid content if any
        except Exception as e:
            print(f"Error with key {key_name}: {e}")
            continue

    return "Failed to generate summary (All keys failed)."

def clean_text_bullets(text):
    """Clean text for Excel cell"""
    if isinstance(text, list):
        clean_items = []
        for item in text:
            clean = str(item).strip()
            if clean:
                clean_items.append(f"• {clean}")
        return '\n'.join(clean_items)

    if not isinstance(text, str):
        return str(text)
        
    text = re.sub(r'<[^>]+>', '', text)
    
    parts = []
    if '•' in text:
         parts = text.split('•')
    else:
         parts = text.split('\n')
        
    clean_items = []
    for p in parts:
        clean = p.strip()
        if clean and not clean.startswith('Stock Ticker'):
             if not clean.startswith('•'):
                 clean = f"• {clean}"
             clean_items.append(clean)
            
    return '\n'.join(clean_items)

def apply_excel_formatting(filename):
    """Apply styling to Excel file"""
    try:
        wb = load_workbook(filename)
        ws = wb.active
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 100
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrapText=True, vertical='top')
        for row in ws.iter_rows(min_row=2, max_col=4):
            for cell in row:
                cell.alignment = Alignment(vertical='top')
        wb.save(filename)
    except Exception as e:
        print(f"Error formating {filename}: {e}")

def main():
    keys = get_gemini_keys()
    print(f"Loaded {len(keys)} Gemini API keys.")
    
    today_str = date.today().strftime("%Y-%m-%d")

    for ticker in TICKERS:
        excel_file = f"{ticker}.xlsx"
        day_num = 1
        
        if os.path.exists(excel_file):
            try:
                existing_df = pd.read_excel(excel_file)
                if not existing_df.empty and 'Day' in existing_df.columns:
                    if 'Date' in existing_df.columns and today_str in existing_df['Date'].astype(str).values:
                        day_num = existing_df[existing_df['Date'].astype(str) == today_str]['Day'].iloc[0]
                        print(f"[{ticker}] Found data for {today_str}. Appending to Day {day_num}.")
                    else:
                        day_num = int(existing_df['Day'].max()) + 1
                        print(f"[{ticker}] New day detected. Starting Day {day_num}.")
            except Exception as e:
                print(f"Error reading {excel_file}: {e}. Starting Day 1.")
        else:
             print(f"[{ticker}] Creating new file. Starting Day 1.")

        print(f"Processing {ticker} (Day {day_num})...")
        news = fetch_news_for_ticker(ticker)
        
        new_rows = []
        for strategy in STRATEGIES:
            print(f"  Generating using strategy: {strategy}...")
            # Retry mechanism is built into generate_summary via key rotation
            summary_raw = generate_summary(ticker, strategy, news, keys)
            summary_clean = clean_text_bullets(summary_raw)
            
            new_rows.append({
                'Date': today_str,
                'Day': day_num,
                'Ticker': ticker,
                'Strategy': strategy,
                'Summary': summary_clean
            })
            time.sleep(0.5)

        new_df = pd.DataFrame(new_rows)
        if os.path.exists(excel_file):
            try:
                existing_df = pd.read_excel(excel_file)
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                combined_df.to_excel(excel_file, index=False)
                print(f"Saved {ticker} results to {excel_file}")
            except:
                 new_df.to_excel(excel_file, index=False)
                 print(f"Overwrote {excel_file} due to error")
        else:
            new_df.to_excel(excel_file, index=False)
            print(f"Created {excel_file}")
            
        apply_excel_formatting(excel_file)

if __name__ == "__main__":
    main()
